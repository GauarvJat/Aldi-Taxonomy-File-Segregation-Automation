import streamlit as st
import pandas as pd
import tempfile
import zipfile
import shutil
import os
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# =====================================================
# CONFIGURATION
# =====================================================

PLATFORM_CONFIG = {
    "AdServer": ["CAMPAIGN", "PLACEMENT", "CREATIVE"],
    "Paid Social": ["CAMPAIGN", "PLACEMENT", "CREATIVE"],
    "Programmatic": ["CAMPAIGN", "PLACEMENT", "PLACEMENTGROUP"],
    "Search": ["CAMPAIGN", "PLACEMENT", "CREATIVE"]
}

INVALID_VALUES = {
    "is missing",
    "invalid",
    "invalid value"
}

RED_FILL = PatternFill(
    start_color="FFFF0000",
    end_color="FFFF0000",
    fill_type="solid"
)

# =====================================================
# HELPER FUNCTIONS
# =====================================================

def parse_filename(filename):
    """
    Example:
    AdServer CAMPAIGN - 16-12-2025.xlsx
    """

    try:

        name_part, date_part = filename.rsplit(" - ", 1)

        date = date_part.replace(".xlsx", "").strip()

        for level in [
            "CAMPAIGN",
            "PLACEMENTGROUP",
            "PLACEMENT",
            "CREATIVE"
        ]:

            if name_part.upper().endswith(level):

                platform = name_part[:-len(level)].strip()

                return platform, level, date

    except:
        pass

    return None, None, None


# =====================================================
# FIND NC RANGE
# =====================================================

def get_nc_indices(df):

    nc_cols = [
        i
        for i, col in enumerate(df.columns)
        if "NC" in str(col).upper()
    ]

    if not nc_cols:
        return None, None

    return nc_cols[0], nc_cols[-1]


# =====================================================
# PROCESS DATAFRAME
# =====================================================

def process_dataframe(df):

    # Rename 2nd Market column

    market_cols = [
        col
        for col in df.columns
        if str(col).strip().lower() == "market"
    ]

    if len(market_cols) > 1:

        df.rename(
            columns={
                market_cols[1]: "Market_MK"
            },
            inplace=True
        )

    start_nc, end_nc = get_nc_indices(df)

    if start_nc is not None:

        for col_index in range(start_nc, end_nc + 1):

            col_name = df.columns[col_index]

            df[col_name] = df[col_name].fillna("Is missing")

    return df


# =====================================================
# EXCEL FORMATTING
# =====================================================

def apply_formatting(file_path):

    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:

        ws = wb[sheet_name]

        headers = [
            cell.value
            for cell in ws[1]
        ]

        nc_cols = [
            i
            for i, header in enumerate(headers)
            if header and "NC" in str(header).upper()
        ]

        if not nc_cols:
            continue

        start_col = min(nc_cols) + 1
        end_col = max(nc_cols) + 1

        for row in ws.iter_rows(
            min_row=2,
            min_col=start_col,
            max_col=end_col
        ):

            for cell in row:

                if cell.value is None:
                    continue

                value = str(cell.value).strip().lower()

                if value in INVALID_VALUES:

                    cell.fill = RED_FILL

    wb.save(file_path)


# =====================================================
# GET ALL MARKETS
# =====================================================

def get_all_markets(files_dict):

    markets = set()

    for file_path in files_dict.values():

        try:

            df = pd.read_excel(
                file_path,
                keep_default_na=False,
                na_values=[""]
            )

            if "Market" in df.columns:

                markets.update(
                    df["Market"]
                    .dropna()
                    .astype(str)
                    .unique()
                )

        except:
            pass

    return sorted(markets)


# =====================================================
# MAIN PROCESSING
# =====================================================

def run_automation(input_folder, output_folder):

    stats = {
        "platforms_processed": 0,
        "files_created": 0,
        "tabs_created": 0
    }

    files_info = {}

    # ----------------------------------
    # Scan Files
    # ----------------------------------

    for file_path in Path(input_folder).glob("*.xlsx"):

        platform, level, date = parse_filename(
            file_path.name
        )

        if not platform:
            continue

        files_info.setdefault(
            platform,
            {
                "date": date,
                "files": {}
            }
        )

        files_info[platform]["files"][level] = file_path

    # ----------------------------------
    # Process Platforms
    # ----------------------------------

    for platform in PLATFORM_CONFIG:

        if platform not in files_info:
            continue

        stats["platforms_processed"] += 1

        platform_files = files_info[platform]["files"]
        date = files_info[platform]["date"]

        markets = get_all_markets(platform_files)

        for market in markets:

            output_file = (
                f"{platform}_{market}_{date}.xlsx"
            )

            output_path = os.path.join(
                output_folder,
                output_file
            )

            created_sheets = []

            with pd.ExcelWriter(
                output_path,
                engine="openpyxl"
            ) as writer:

                for level in PLATFORM_CONFIG[platform]:

                    if level not in platform_files:
                        continue

                    df = pd.read_excel(
                        platform_files[level],
                        keep_default_na=False,
                        na_values=[""]
                    )

                    if df.empty:
                        continue

                    df = process_dataframe(df)

                    if "Market" not in df.columns:
                        continue

                    market_df = df[
                        df["Market"].astype(str)
                        == str(market)
                    ]

                    if market_df.empty:
                        continue

                    market_df.to_excel(
                        writer,
                        sheet_name=level,
                        index=False
                    )

                    created_sheets.append(level)

                    stats["tabs_created"] += 1

            if created_sheets:

                apply_formatting(output_path)

                stats["files_created"] += 1

            else:

                os.remove(output_path)

    return stats


# =====================================================
# STREAMLIT UI
# =====================================================

st.set_page_config(
    page_title="Aldi Data Segregation",
    layout="wide"
)

st.title("📊 Aldi Data Segregation Tool")

st.markdown(
    """
Upload all source Excel files and click **Run Automation**.
The tool will generate Market-level files and provide a ZIP download.
"""
)

uploaded_files = st.file_uploader(
    "Upload Aldi Source Files",
    type=["xlsx"],
    accept_multiple_files=True
)

if st.button("🚀 Run Automation"):

    if not uploaded_files:

        st.error("Please upload files.")

    else:

        with tempfile.TemporaryDirectory() as temp_dir:

            input_folder = os.path.join(
                temp_dir,
                "input"
            )

            output_folder = os.path.join(
                temp_dir,
                "output"
            )

            os.makedirs(input_folder)
            os.makedirs(output_folder)

            # Save uploads

            for uploaded_file in uploaded_files:

                save_path = os.path.join(
                    input_folder,
                    uploaded_file.name
                )

                with open(save_path, "wb") as f:

                    f.write(uploaded_file.getbuffer())

            progress = st.progress(0)

            stats = run_automation(
                input_folder,
                output_folder
            )

            progress.progress(100)

            # Create ZIP

            zip_path = os.path.join(
                temp_dir,
                "Aldi_Output.zip"
            )

            with zipfile.ZipFile(
                zip_path,
                "w",
                zipfile.ZIP_DEFLATED
            ) as zipf:

                for file in Path(output_folder).glob("*.xlsx"):

                    zipf.write(
                        file,
                        arcname=file.name
                    )

            st.success("Automation Completed Successfully")

            col1, col2, col3 = st.columns(3)

            col1.metric(
                "Platforms",
                stats["platforms_processed"]
            )

            col2.metric(
                "Files Created",
                stats["files_created"]
            )

            col3.metric(
                "Tabs Created",
                stats["tabs_created"]
            )

            with open(zip_path, "rb") as f:

                st.download_button(
                    label="📥 Download Output ZIP",
                    data=f,
                    file_name="Aldi_Output.zip",
                    mime="application/zip"
                )