from pathlib import Path

app_code = r'''
# Aldi Streamlit Data Segregation Tool
# Production-ready Streamlit app

import io
import zipfile
from pathlib import PurePath
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

PLATFORM_CONFIG = {
    "AdServer": ["CAMPAIGN", "PLACEMENT", "CREATIVE"],
    "Paid Social": ["CAMPAIGN", "PLACEMENT", "CREATIVE"],
    "Programmatic": ["CAMPAIGN", "PLACEMENT", "PLACEMENTGROUP"],
    "Search": ["CAMPAIGN", "PLACEMENT", "CREATIVE"]
}

PLATFORM_ALIASES = {
    "adserver": "AdServer",
    "paid social": "Paid Social",
    "paidsocial": "Paid Social",
    "programmatic": "Programmatic",
    "search": "Search",
}

INVALID_VALUES = {"is missing", "invalid", "invalid value"}

def parse_filename(filename):
    stem = PurePath(filename).stem.strip()
    try:
        name_part, date = stem.rsplit(" - ", 1)
    except ValueError:
        return None, None, None

    for level in ["PLACEMENTGROUP", "CAMPAIGN", "PLACEMENT", "CREATIVE"]:
        if name_part.upper().endswith(level):
            platform = name_part[:-len(level)].strip()
            return platform, level, date
    return None, None, None

def process_dataframe(df):
    market_cols = [c for c in df.columns if str(c).lower() == "market"]
    if len(market_cols) > 1:
        df.rename(columns={market_cols[1]: "Market_MK"}, inplace=True)

    nc_cols = [i for i, c in enumerate(df.columns) if "NC" in str(c).upper()]
    if nc_cols:
        start, end = min(nc_cols), max(nc_cols)
        for idx in range(start, end + 1):
            df.iloc[:, idx] = df.iloc[:, idx].replace("", pd.NA).fillna("Is missing")
    return df

def apply_formatting_to_workbook(wb):
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    for ws in wb.worksheets:
        headers = [c.value for c in ws[1]]
        nc_cols = [i for i, h in enumerate(headers) if h and "NC" in str(h).upper()]
        if not nc_cols:
            continue

        start_col = min(nc_cols) + 1
        end_col = max(nc_cols) + 1

        for row in ws.iter_rows(min_row=2, min_col=start_col, max_col=end_col):
            for cell in row:
                if cell.value and str(cell.value).strip().lower() in INVALID_VALUES:
                    cell.fill = red_fill
    return wb

def run_segregation(uploaded_files, progress_cb=None):
    inputs = {}
    stats = {"platforms_processed":0,"files_created":0,"tabs_created":0,"skipped":[]}

    for uf in uploaded_files:
        platform_raw, level, date = parse_filename(uf.name)
        if not platform_raw:
            stats["skipped"].append(f"{uf.name} - invalid filename")
            continue

        platform = PLATFORM_ALIASES.get(platform_raw.lower())
        if not platform:
            stats["skipped"].append(f"{uf.name} - unknown platform")
            continue

        inputs.setdefault(platform, {}).setdefault(date, {})[level] = uf

    output_files = {}
    total = sum(len(v) for v in inputs.values()) or 1
    done = 0

    for platform, dates in inputs.items():
        stats["platforms_processed"] += 1

        for date, levels in dates.items():
            done += 1
            if progress_cb:
                progress_cb(done / total)

            market_data = {}

            for level, uf in levels.items():
                uf.seek(0)
                df = pd.read_excel(io.BytesIO(uf.read()), keep_default_na=False, na_values=[""])

                if df.empty or "Market" not in df.columns:
                    continue

                df = process_dataframe(df)

                for market, group in df.groupby("Market"):
                    market_data.setdefault(str(market), {})[level] = group

            for market, level_data in market_data.items():
                buf = io.BytesIO()

                with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                    for level in PLATFORM_CONFIG[platform]:
                        if level in level_data:
                            level_data[level].to_excel(writer, sheet_name=level, index=False)
                            stats["tabs_created"] += 1

                buf.seek(0)
                wb = load_workbook(buf)
                wb = apply_formatting_to_workbook(wb)

                final = io.BytesIO()
                wb.save(final)

                output_files[f"{platform}_{market}_{date}.xlsx"] = final.getvalue()
                stats["files_created"] += 1

    return output_files, stats

def build_zip(output_files):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname, data in output_files.items():
            zf.writestr(fname, data)
    return buf.getvalue()

st.set_page_config(page_title="ALDI Data Segregation", page_icon="🛒", layout="centered")

st.title("🛒 ALDI Data Segregation Tool")
st.write("Upload source Excel files, run segregation, and download the ZIP output.")

uploaded_files = st.file_uploader(
    "Upload Aldi source files",
    type=["xlsx"],
    accept_multiple_files=True
)

if st.button("Run Segregation", disabled=not uploaded_files):
    progress = st.progress(0)

    def cb(x):
        progress.progress(min(int(x*100), 100))

    output_files, stats = run_segregation(uploaded_files, cb)

    st.success("Processing completed")

    c1, c2, c3 = st.columns(3)
    c1.metric("Platforms", stats["platforms_processed"])
    c2.metric("Files", stats["files_created"])
    c3.metric("Tabs", stats["tabs_created"])

    if stats["skipped"]:
        with st.expander("Skipped Items"):
            for s in stats["skipped"]:
                st.write("-", s)

    zip_bytes = build_zip(output_files)

    st.download_button(
        "Download Output ZIP",
        data=zip_bytes,
        file_name="Aldi_Segregated_Output.zip",
        mime="application/zip"
    )
'''

path = "/mnt/data/app.py"
Path(path).write_text(app_code, encoding="utf-8")

print(path)
